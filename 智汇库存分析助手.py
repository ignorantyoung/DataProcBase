import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os

class ColumnSelector:
    def __init__(self, columns, data):
        self.window = tk.Toplevel()
        self.window.title("选择要保存的列")
        self.window.geometry("1200x800")
        
        # 设置默认字体大小
        default_font = ("TkDefaultFont", 14)  # 增大默认字体
        self.window.option_add("*Font", default_font)
        self.data = data
        self.columns = columns
        
        # 创建主框架
        self.main_frame = ttk.Frame(self.window)
        self.main_frame.pack(fill="both", expand=True, padx=5, pady=3)  # 减小边距
        
        # 左侧框架 - 列选择
        self.left_frame = ttk.LabelFrame(self.main_frame, text="选择要保存的列", padding="5")
        self.left_frame.pack(side="left", fill="both", expand=True, padx=3)
        
        # 创建按钮框架
        self.button_control_frame = ttk.Frame(self.left_frame)
        self.button_control_frame.pack(fill="x", padx=5, pady=5)
        
        # 添加全选/全不选按钮
        ttk.Button(self.button_control_frame, text="全选", command=self.select_all).pack(side="left", padx=5)
        ttk.Button(self.button_control_frame, text="全不选", command=self.deselect_all).pack(side="left", padx=5)
        
        # 创建滚动框架
        self.scroll_frame = ttk.Frame(self.left_frame)
        self.scroll_frame.pack(fill="both", expand=True)
        
        # 添加滚动条
        self.scrollbar = ttk.Scrollbar(self.scroll_frame)
        self.scrollbar.pack(side="right", fill="y")
        
        # 创建画布
        self.canvas = tk.Canvas(self.scroll_frame)
        self.canvas.pack(side="left", fill="both", expand=True)
        
        # 配置滚动条
        self.scrollbar.config(command=self.canvas.yview)
        self.canvas.config(yscrollcommand=self.scrollbar.set)
        
        # 创建复选框容器
        self.checkbox_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.checkbox_frame, anchor="nw")
        
        # 创建列选择变量和复选框
        self.column_vars = {}
        for col in columns:
            var = tk.BooleanVar(value=True)
            self.column_vars[col] = var
            ttk.Checkbutton(self.checkbox_frame, text=col, variable=var, command=self.update_preview).pack(anchor="w")
        
        # 绑定画布滚动事件
        self.checkbox_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        
        # 中间框架 - 新列配置
        self.middle_frame = ttk.LabelFrame(self.main_frame, text="新列配置", padding="5")
        self.middle_frame.pack(side="left", fill="both", expand=True, padx=3)
        
        # 添加新列复选框
        self.add_new_column_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(self.middle_frame, text="添加计算列", variable=self.add_new_column_var, command=self.toggle_new_column_config).pack(anchor="w", pady=3)
        
        # 新列配置框架
        self.new_column_config_frame = ttk.Frame(self.middle_frame)
        
        # 新列名称输入
        ttk.Label(self.new_column_config_frame, text="新列名称:").pack(anchor="w", pady=3)
        self.new_column_name = ttk.Entry(self.new_column_config_frame)
        self.new_column_name.pack(fill="x", pady=3)
        
        # 计算方式选择
        self.calc_mode_frame = ttk.Frame(self.new_column_config_frame)
        self.calc_mode_frame.pack(fill="x", pady=3)
        
        # 移除基础计算模式，只保留高级公式模式
        self.calc_mode = tk.StringVar(value="advanced")
        
        # 初始化简单计算模式的组件
        self.column_combobox = ttk.Combobox(self.calc_mode_frame)
        self.operator_combobox = ttk.Combobox(self.calc_mode_frame)
        self.column_combobox2 = ttk.Combobox(self.calc_mode_frame)
        self.multiplier = ttk.Entry(self.calc_mode_frame)
        
        # 高级公式框架
        self.advanced_calc_frame = ttk.Frame(self.new_column_config_frame)
        
        # 添加预设公式选择
        preset_frame = ttk.LabelFrame(self.advanced_calc_frame, text="预设公式", padding=5)
        preset_frame.pack(fill="x", pady=5)
        
        self.preset_formula = ttk.Combobox(preset_frame, state="readonly")
        self.preset_formula["values"] = [
            "自定义公式",
            "库存天数(不含在途) = 运营云仓可用数/30天发货量 * 30",
            "库存天数(含在途) = (运营云仓可用数 + 采购在途)/30天发货量 * 30"
        ]
        self.preset_formula.current(0)
        self.preset_formula.pack(fill="x", pady=5)
        self.preset_formula.bind("<<ComboboxSelected>>", self.on_preset_formula_selected)
        
        # 简化的公式说明框架
        formula_help_frame = ttk.LabelFrame(self.advanced_calc_frame, text="公式使用说明", padding=5)
        formula_help_frame.pack(fill="x", pady=5)
        
        # 简化的公式使用说明
        help_text = """公式使用说明：
1. 使用英文方括号[列名]引用列
2. 支持运算符：+, -, *, /
3. 支持英文括号()设置优先级
4. 示例：[销售额] - [成本]"""
        ttk.Label(formula_help_frame, text=help_text, wraplength=400, justify="left").pack(pady=5)
        
        # 公式输入框（调整高度）
        ttk.Label(self.advanced_calc_frame, text="计算公式:").pack(anchor="w", pady=2)
        formula_frame = ttk.Frame(self.advanced_calc_frame)
        formula_frame.pack(fill="x", pady=2)
        
        self.formula_entry = tk.Text(formula_frame, height=2, wrap=tk.WORD, font=("TkDefaultFont", 11))
        self.formula_entry.pack(fill="x", side="left", expand=True)
        
        # 为公式输入框添加滚动条
        formula_scrollbar = ttk.Scrollbar(formula_frame, orient="vertical", command=self.formula_entry.yview)
        formula_scrollbar.pack(side="right", fill="y")
        self.formula_entry.configure(yscrollcommand=formula_scrollbar.set)
        
        # 可用列列表（只显示已选择的列）
        ttk.Label(self.advanced_calc_frame, text="可用列:").pack(anchor="w", pady=5)
        columns_frame = ttk.Frame(self.advanced_calc_frame)
        columns_frame.pack(fill="x", pady=5)
        
        self.columns_listbox = tk.Listbox(columns_frame, height=5)
        self.columns_listbox.pack(fill="x", side="left", expand=True)
        
        # 为列表框添加滚动条
        columns_scrollbar = ttk.Scrollbar(columns_frame, orient="vertical", command=self.columns_listbox.yview)
        columns_scrollbar.pack(side="right", fill="y")
        self.columns_listbox.configure(yscrollcommand=columns_scrollbar.set)
        
        # 公式帮助文本
        help_text = "支持的运算符: +, -, *, /, (, )\n双击列表中的列名添加到公式中"
        ttk.Label(self.advanced_calc_frame, text=help_text).pack(anchor="w", pady=5)
        
        # 绑定双击事件
        self.columns_listbox.bind('<Double-Button-1>', self.insert_column_to_formula)
        
        # 添加计算按钮
        ttk.Button(self.new_column_config_frame, text="添加计算列", command=self.add_calculated_column).pack(pady=10)
        
        # 默认显示高级计算框架
        self.advanced_calc_frame.pack(fill="x", pady=5)
        
        # 右侧框架 - 预览
        self.right_frame = ttk.LabelFrame(self.main_frame, text="数据预览", padding="10")
        self.right_frame.pack(side="right", fill="both", expand=True, padx=5)
        
        # 创建预览表格
        self.preview_tree = ttk.Treeview(self.right_frame, style="Custom.Treeview")
        self.preview_tree.pack(fill="both", expand=True)
        
        # 设置预览表格样式
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=25, font=("TkDefaultFont", 11))  # 减小表格字体和行高
        style.configure("Custom.Treeview.Heading", font=("TkDefaultFont", 12))  # 减小表头字体
        
        # 预览表格的滚动条
        scrollbar = ttk.Scrollbar(self.right_frame, orient="vertical", command=self.preview_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.preview_tree.configure(yscrollcommand=scrollbar.set)
        
        # 确认和取消按钮
        self.button_frame = ttk.Frame(self.window)
        self.button_frame.pack(fill="x", padx=10, pady=5)
        ttk.Button(self.button_frame, text="确认", command=self.confirm).pack(side="right", padx=5)
        ttk.Button(self.button_frame, text="取消", command=self.cancel).pack(side="right", padx=5)
        
        self.selected_columns = []
        self.confirmed = False
        
        # 初始化预览
        self.update_preview()
        
        # 初始隐藏新列配置
        self.new_column_config_frame.pack_forget()
    
    def toggle_new_column_config(self):
        if self.add_new_column_var.get():
            self.new_column_config_frame.pack(fill="both", expand=True)
        else:
            self.new_column_config_frame.pack_forget()
    

    def validate_formula(self, formula):
        # 检查中文括号
        if '（' in formula or '）' in formula:
            messagebox.showwarning("警告", "请使用英文括号 () 而不是中文括号（）")
            return False
            
        # 检查方括号是否成对出现
        if formula.count('[') != formula.count(']'):
            messagebox.showwarning("警告", "公式中的方括号不匹配，请检查")
            return False
            
        # 检查括号是否成对出现
        if formula.count('(') != formula.count(')'):
            messagebox.showwarning("警告", "公式中的括号不匹配，请检查")
            return False
            
        # 检查是否包含非法字符
        import re
        # 允许中文字符作为列名，但确保它们在方括号内
        column_pattern = r'\[(.*?)\]'
        columns = re.findall(column_pattern, formula)
        
        # 检查引用的列名是否存在
        for col in columns:
            if col not in self.columns:
                messagebox.showwarning("警告", f"公式中引用了不存在的列：{col}")
                return False
        
        # 移除所有列引用，检查剩余部分是否只包含合法字符
        formula_without_columns = re.sub(column_pattern, '', formula)
        allowed_chars = r'[\(\)\+\-\*\/\s\d\.]'
        illegal_chars = re.sub(allowed_chars, '', formula_without_columns)
        illegal_chars = ''.join(set(illegal_chars))
        
        if illegal_chars:
            messagebox.showwarning("警告", f"公式中包含非法字符：{illegal_chars}\n只允许使用：数字、运算符(+-*/)、英文括号()")
            return False
            
        return True
        
    def validate_processed_formula(self, formula):
        # 检查是否包含危险的Python内置函数或属性
        dangerous_terms = ['__', 'eval', 'exec', 'import', 'open', 'os', 'sys']
        for term in dangerous_terms:
            if term in formula:
                messagebox.showwarning("警告", f"公式中包含不允许使用的关键字：{term}")
                return False
                
        # 检查列引用是否有效
        import re
        column_refs = re.findall(r"self\.data\['([^']+)'\]", formula)
        for col in column_refs:
            if col not in self.columns:
                messagebox.showwarning("警告", f"公式中引用了不存在的列：{col}")
                return False
                
        return True
    
    def insert_column_to_formula(self, event):
        selection = self.columns_listbox.curselection()
        if selection:
            column = self.columns_listbox.get(selection[0])
            current_pos = self.formula_entry.index(tk.INSERT)
            self.formula_entry.insert(current_pos, f"[{column}]")
    
    def add_calculated_column(self):
        try:
            import re  # 确保re模块在函数开始时就被导入
            import numpy as np  # 直接导入numpy
            new_column_name = self.new_column_name.get().strip()
            if not new_column_name:
                messagebox.showwarning("警告", "请输入新列名称")
                return
            
            # 高级公式模式
            formula = self.formula_entry.get("1.0", tk.END).strip()
            if not formula:
                messagebox.showwarning("警告", "请输入计算公式")
                return
            
            # 验证公式格式
            if not self.validate_formula(formula):
                return
            
            # 替换公式中的列名
            processed_formula = formula
            for col in self.columns:
                # 使用正则表达式确保完整匹配列名
                pattern = re.escape(f"[{col}]")
                processed_formula = re.sub(pattern, f"self.data['{col}']", processed_formula)
            
            # 验证处理后的公式
            if not self.validate_processed_formula(processed_formula):
                return
            
            try:
                # 计算结果
                # 使用更安全的方式计算公式
                namespace = {'self': self, 'pd': pd, 'np': np}
                result = eval(processed_formula, {'__builtins__': {}}, namespace)
                if isinstance(result, pd.Series):
                    result = result.values
                elif isinstance(result, (int, float)):
                    # 如果结果是单个数值，扩展为与数据行数相同的数组
                    result = np.full(len(self.data), result)
            except NameError as e:
                messagebox.showerror("错误", f"公式中包含未定义的变量：\n{str(e)}\n请检查列名是否正确")
                return
            except TypeError as e:
                messagebox.showerror("错误", f"公式中的数据类型不匹配：\n{str(e)}\n请检查运算是否合法")
                return
            except ZeroDivisionError:
                messagebox.showerror("错误", "公式中出现除以零的操作，请检查")
                return
            except Exception as e:
                messagebox.showerror("错误", f"计算公式错误：\n{str(e)}\n请检查公式格式是否正确")
                return
            
            # 添加新列到数据中
            try:
                self.data[new_column_name] = result
            except Exception as e:
                messagebox.showerror("错误", f"添加新列失败：\n{str(e)}\n请检查数据是否合法")
                return
            
            # 更新列选择界面
            var = tk.BooleanVar(value=True)
            self.column_vars[new_column_name] = var
            ttk.Checkbutton(self.checkbox_frame, text=new_column_name, variable=var, command=self.update_preview).pack(anchor="w")
            
            # 更新预览
            self.update_preview()
            
            # 清空输入
            self.new_column_name.delete(0, tk.END)
            self.formula_entry.delete("1.0", tk.END)
            
            messagebox.showinfo("成功", f"已添加新列: {new_column_name}")
            
        except Exception as e:
            messagebox.showerror("错误", f"计算过程中出错：\n{str(e)}")
    
    def update_preview(self):
        # 清空现有预览
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        # 获取选中的列（保持原始顺序）
        selected_cols = [col for col in self.columns if self.column_vars[col].get()]
        if not selected_cols:
            return
        
        # 更新预览表格的列
        self.preview_tree["columns"] = selected_cols
        self.preview_tree["show"] = "headings"
        
        for col in selected_cols:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=100)
        
        # 显示前5行数据
        preview_data = self.data[selected_cols].head()
        for _, row in preview_data.iterrows():
            self.preview_tree.insert("", tk.END, values=list(row))
        
        # 更新可用列列表
        self.columns_listbox.delete(0, tk.END)
        for col in selected_cols:
            self.columns_listbox.insert(tk.END, col)
        
        # 更新简单计算模式的下拉列表
        self.column_combobox["values"] = selected_cols
        self.column_combobox2["values"] = selected_cols
    
    def confirm(self):
        self.selected_columns = [col for col, var in self.column_vars.items() if var.get()]
        if not self.selected_columns:
            messagebox.showwarning("警告", "请至少选择一列")
            return
        self.confirmed = True
        self.window.destroy()
    
    def cancel(self):
        self.confirmed = False
        self.window.destroy()

    def select_all(self):
        for var in self.column_vars.values():
            var.set(True)
        self.update_preview()
    
    def deselect_all(self):
        for var in self.column_vars.values():
            var.set(False)
        self.update_preview()
    
    def on_frame_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    
    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas.find_all()[0], width=event.width)

    def on_preset_formula_selected(self, event):
        selected = self.preset_formula.get()
        if selected == "自定义公式":
            self.formula_entry.delete("1.0", tk.END)
            return
            
        if selected.startswith("库存天数(不含在途)"):
            formula = "[运营云仓可用数] / [30天发货量] * 30"
            self.new_column_name.delete(0, tk.END)
            self.new_column_name.insert(0, "库存天数(不含在途)")
        elif selected.startswith("库存天数(含在途)"):
            formula = "([运营云仓可用数] + [采购在途]) / [30天发货量] * 30"
            self.new_column_name.delete(0, tk.END)
            self.new_column_name.insert(0, "库存天数(含在途)")
            
        self.formula_entry.delete("1.0", tk.END)
        self.formula_entry.insert("1.0", formula)
    
    def insert_column_to_formula(self, event):
        selection = self.columns_listbox.curselection()
        if selection:
            column = self.columns_listbox.get(selection[0])
            current_pos = self.formula_entry.index(tk.INSERT)
            self.formula_entry.insert(current_pos, f"[{column}]")
    
    def add_calculated_column(self):
        try:
            import re  # 确保re模块在函数开始时就被导入
            import numpy as np  # 直接导入numpy
            new_column_name = self.new_column_name.get().strip()
            if not new_column_name:
                messagebox.showwarning("警告", "请输入新列名称")
                return
            
            # 高级公式模式
            formula = self.formula_entry.get("1.0", tk.END).strip()
            if not formula:
                messagebox.showwarning("警告", "请输入计算公式")
                return
            
            # 验证公式格式
            if not self.validate_formula(formula):
                return
            
            # 替换公式中的列名
            processed_formula = formula
            for col in self.columns:
                # 使用正则表达式确保完整匹配列名
                pattern = re.escape(f"[{col}]")
                processed_formula = re.sub(pattern, f"self.data['{col}']", processed_formula)
            
            # 验证处理后的公式
            if not self.validate_processed_formula(processed_formula):
                return
            
            try:
                # 计算结果
                # 使用更安全的方式计算公式
                namespace = {'self': self, 'pd': pd, 'np': np}
                result = eval(processed_formula, {'__builtins__': {}}, namespace)
                if isinstance(result, pd.Series):
                    result = result.values
                elif isinstance(result, (int, float)):
                    # 如果结果是单个数值，扩展为与数据行数相同的数组
                    result = np.full(len(self.data), result)
            except NameError as e:
                messagebox.showerror("错误", f"公式中包含未定义的变量：\n{str(e)}\n请检查列名是否正确")
                return
            except TypeError as e:
                messagebox.showerror("错误", f"公式中的数据类型不匹配：\n{str(e)}\n请检查运算是否合法")
                return
            except ZeroDivisionError:
                messagebox.showerror("错误", "公式中出现除以零的操作，请检查")
                return
            except Exception as e:
                messagebox.showerror("错误", f"计算公式错误：\n{str(e)}\n请检查公式格式是否正确")
                return
            
            # 添加新列到数据中
            try:
                self.data[new_column_name] = result
            except Exception as e:
                messagebox.showerror("错误", f"添加新列失败：\n{str(e)}\n请检查数据是否合法")
                return
            
            # 更新列选择界面
            var = tk.BooleanVar(value=True)
            self.column_vars[new_column_name] = var
            ttk.Checkbutton(self.checkbox_frame, text=new_column_name, variable=var, command=self.update_preview).pack(anchor="w")
            
            # 更新预览
            self.update_preview()
            
            # 清空输入
            self.new_column_name.delete(0, tk.END)
            self.formula_entry.delete("1.0", tk.END)
            
            messagebox.showinfo("成功", f"已添加新列: {new_column_name}")
            
        except Exception as e:
            messagebox.showerror("错误", f"计算过程中出错：\n{str(e)}")
    
    def update_preview(self):
        # 清空现有预览
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        # 获取选中的列（保持原始顺序）
        selected_cols = [col for col in self.columns if self.column_vars[col].get()]
        if not selected_cols:
            return
        
        # 更新预览表格的列
        self.preview_tree["columns"] = selected_cols
        self.preview_tree["show"] = "headings"
        
        for col in selected_cols:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=100)
        
        # 显示前5行数据
        preview_data = self.data[selected_cols].head()
        for _, row in preview_data.iterrows():
            self.preview_tree.insert("", tk.END, values=list(row))
        
        # 更新可用列列表
        self.columns_listbox.delete(0, tk.END)
        for col in selected_cols:
            self.columns_listbox.insert(tk.END, col)
        
        # 更新简单计算模式的下拉列表
        self.column_combobox["values"] = selected_cols
        self.column_combobox2["values"] = selected_cols
    
    def confirm(self):
        self.selected_columns = [col for col, var in self.column_vars.items() if var.get()]
        if not self.selected_columns:
            messagebox.showwarning("警告", "请至少选择一列")
            return
        self.confirmed = True
        self.window.destroy()
    
    def cancel(self):
        self.confirmed = False
        self.window.destroy()

    def select_all(self):
        for var in self.column_vars.values():
            var.set(True)
        self.update_preview()
    
    def deselect_all(self):
        for var in self.column_vars.values():
            var.set(False)
        self.update_preview()
    
    def on_frame_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    
    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas.find_all()[0], width=event.width)

def process_inventory(file_path):
    try:
        # 验证文件路径
        if not os.path.exists(file_path):
            messagebox.showerror("错误", "选择的文件不存在，请检查文件路径")
            return
            
        # 读取原始数据
        try:
            data = pd.read_excel(file_path)
        except Exception as e:
            messagebox.showerror("错误", f"读取Excel文件失败：\n{str(e)}\n请确保文件格式正确且未被其他程序占用")
            return
            
        # 处理空值，将所有空值替换为0
        data = data.fillna(0)
        
        # 将所有数值列四舍五入为整数
        numeric_columns = data.select_dtypes(include=['float64', 'int64']).columns
        for col in numeric_columns:
            data[col] = data[col].round(0)
        
        # 添加预设公式计算
        try:
            # 公式1：运营云仓可用数/30天发货量 * 30
            if '运营云仓可用数' in data.columns and '30天发货量' in data.columns:
                # 添加除数为0的检查
                data['库存天数(不含在途)'] = (data['运营云仓可用数'] / data['30天发货量'].replace(0, float('nan')) * 30).round(0).astype('Int64')
            
            # 公式2：(运营云仓可用数 + 采购在途) /30天发货量 * 30
            if '运营云仓可用数' in data.columns and '采购在途' in data.columns and '30天发货量' in data.columns:
                # 添加除数为0的检查
                data['库存天数(含在途)'] = ((data['运营云仓可用数'] + data['采购在途']) / data['30天发货量'].replace(0, float('nan')) * 30).round(0).astype('Int64')
        except Exception as e:
            messagebox.showwarning("警告", f"计算预设公式时出错：\n{str(e)}")
        
        # 显示列选择界面
        selector = ColumnSelector(data.columns, data)
        selector.window.wait_window()
        
        if not selector.confirmed:
            return
    
        # 筛选选中的列
        filtered_data = data[selector.selected_columns]
        
        # 创建一个新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        
        # 将所有数值列转换为整数
        for col in filtered_data.columns:
            try:
                filtered_data[col] = filtered_data[col].round(0).astype('Int64')
            except:
                continue
                
        # 将DataFrame写入工作表
        for r in dataframe_to_rows(filtered_data, index=False, header=True):
            ws.append(r)
            
        # 设置数值列的格式为整数
        from openpyxl.styles import numbers
        for col in range(1, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                try:
                    float(cell.value)  # 检查是否为数值
                    cell.number_format = numbers.FORMAT_NUMBER  # 设置为无小数位的数值格式
                except (ValueError, TypeError):
                    continue  # 如果不是数值，保持原格式

        # 应用条件格式化
        from openpyxl.styles import PatternFill
        
        # 定义颜色
        deep_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色
        red = PatternFill(start_color='FF60D0', end_color='FF60D0', fill_type='solid')       # 玫瑰红
        light_yellow = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')    # 浅黄色
        light_blue = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')      # 天蓝色
        
        # 获取计算列的列号
        calc_columns = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header in ['库存天数(不含在途)', '库存天数(含在途)'] or header.startswith('计算列'):
                calc_columns.append(col)

        # 遍历所有数据单元格，只对计算列应用颜色标记
        for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
            for col in calc_columns:  # 只遍历计算列
                cell = ws.cell(row=row, column=col)
                try:
                    value = float(cell.value)
                    if value < 0:
                        cell.fill = deep_red
                    elif 0 <= value < 10:
                        cell.fill = red
                    elif 10 <= value < 30:
                        cell.fill = light_yellow
                    elif 30 <= value < 60:
                        cell.fill = light_blue
                except (ValueError, TypeError):
                    continue  # 如果单元格值不能转换为数字，跳过
        
        # 创建data文件夹（如果不存在）
        data_dir = "data"
        try:
            if not os.path.exists(data_dir):
                os.makedirs(data_dir)
        except Exception as e:
            messagebox.showerror("错误", f"创建data文件夹失败：\n{str(e)}\n请确保程序有足够的权限")
            return
        
        # 保存工作簿到data文件夹中
        try:
            output_file = os.path.join(data_dir, "处理后的数据_" + os.path.basename(file_path))
            # 如果文件已存在，先尝试删除
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                except Exception as e:
                    messagebox.showerror("错误", f"无法删除已存在的文件：\n{str(e)}\n请确保文件未被其他程序占用")
                    return
            
            # 尝试保存文件
            wb.save(output_file)
            messagebox.showinfo("完成", f"数据已处理完成并保存到: {output_file}")
        except Exception as e:
            messagebox.showerror("错误", f"保存文件时出错：\n{str(e)}\n请确保有足够的权限和磁盘空间")
    except Exception as e:
        messagebox.showerror("错误", f"处理文件时出错：\n{str(e)}")

if __name__ == "__main__":
    try:
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        
        # 显示欢迎信息和操作提示
        messagebox.showinfo("欢迎", "请选择需要处理的Excel文件\n处理后的文件将保存在data文件夹中")
        
        file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel files", "*.xlsx *.xls")])
        
        if file_path:
            try:
                process_inventory(file_path)
            except Exception as e:
                messagebox.showerror("错误", f"处理文件时出错：\n{str(e)}")
        else:
            messagebox.showwarning("提示", "未选择文件，程序将退出")
    except Exception as e:
        messagebox.showerror("错误", f"程序运行出错：\n{str(e)}")
